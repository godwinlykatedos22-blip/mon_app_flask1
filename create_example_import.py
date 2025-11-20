import sys, os
sys.path.insert(0, os.path.abspath(os.getcwd()))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# Créer le workbook
wb = Workbook()
ws = wb.active
ws.title = "Élèves"

# Définir les largeurs de colonnes
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 15

# Styles
header_fill = PatternFill(start_color="003F7F", end_color="003F7F", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# En-têtes
headers = ["Nom", "Prénom", "Date de naissance", "Classe", "Prénom Parent/Tuteur", "Nom Parent/Tuteur", "Téléphone Parent (E.164)", "WhatsApp (Oui/Non)"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Données d'exemple - Classe 6ème A
data = [
    ["Dupont", "Jean", date(2008, 5, 15), "6ème A", "Marie", "Dupont", "+33612345678", "Oui"],
    ["Dupont", "Sophie", date(2010, 3, 10), "6ème A", "Marie", "Dupont", "+33612345678", "Oui"],
    ["Martin", "Alice", date(2009, 1, 20), "6ème A", "Pierre", "Martin", "+33687654321", "Non"],
    ["Bernard", "Luc", date(2008, 11, 5), "6ème A", "Anne", "Bernard", "+33698765432", "Oui"],
    ["Lefevre", "Emma", date(2009, 7, 22), "6ème A", "Marc", "Lefevre", "+33645123789", "Oui"],
    ["Lefevre", "Thomas", date(2010, 9, 12), "6ème A", "Marc", "Lefevre", "+33645123789", "Oui"],
    ["Moreau", "Julie", date(2008, 4, 8), "6ème A", "Francoise", "Moreau", "+33756234891", "Non"],
    ["Girard", "Nicolas", date(2009, 6, 30), "6ème A", "Laurent", "Girard", "+33789456123", "Oui"],
]

# Ajouter les données
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        
        # Formatage spécifique par colonne
        if col_num == 3:  # Date de naissance
            cell.number_format = 'YYYY-MM-DD'
            cell.alignment = Alignment(horizontal='center')
        elif col_num in [4, 8]:  # Classe et WhatsApp
            cell.alignment = Alignment(horizontal='center')

# Feuille Classes disponibles
ws2 = wb.create_sheet("Classes disponibles")
ws2.column_dimensions['A'].width = 20

cell = ws2['A1']
cell.value = "Nom de la classe"
cell.fill = header_fill
cell.font = header_font
cell.border = border

classes = ["6ème A", "6ème B", "5ème A", "5ème B", "4ème A", "4ème B", "3ème A", "3ème B", "2nde AB", "1ère AB", "Tle AB"]
for idx, classe_name in enumerate(classes, 2):
    cell = ws2.cell(row=idx, column=1)
    cell.value = classe_name
    cell.border = border
    cell.alignment = Alignment(horizontal='left')

# Feuille Instructions
ws3 = wb.create_sheet("Instructions")
ws3.column_dimensions['A'].width = 100

title_cell = ws3['A1']
title_cell.value = "INSTRUCTIONS D'IMPORT - ÉLÈVES ET PARENTS"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="003F7F", end_color="003F7F", fill_type="solid")
title_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws3.row_dimensions[1].height = 30

instructions = [
    "",
    "📋 REMPLISSAGE DU FICHIER",
    "",
    "Colonne A - Nom (OBLIGATOIRE)",
    "  • Nom de famille de l'élève",
    "  • Exemple : Dupont",
    "",
    "Colonne B - Prénom (OBLIGATOIRE)",
    "  • Prénom de l'élève",
    "  • Exemple : Jean",
    "",
    "Colonne C - Date de naissance (OPTIONNEL)",
    "  • Format : YYYY-MM-DD (exemple : 2008-05-15)",
    "  • Peut être laissée vide",
    "",
    "Colonne D - Classe (OBLIGATOIRE)",
    "  • Doit correspondre exactement à une classe existante",
    "  • Voir la feuille 'Classes disponibles'",
    "  • Exemple : 6ème A",
    "",
    "Colonne E - Prénom Parent/Tuteur (OPTIONNEL)",
    "  • Prénom du parent ou tuteur",
    "  • Exemple : Marie",
    "",
    "Colonne F - Nom Parent/Tuteur (OPTIONNEL)",
    "  • Nom du parent ou tuteur",
    "  • Exemple : Dupont",
    "",
    "Colonne G - Téléphone Parent (OPTIONNEL)",
    "  • Format E.164 : commence par + suivi du code pays",
    "  • Exemple pour France : +33612345678 (remplacer le 0 par 33)",
    "  • Important : Un parent est reconnu par son téléphone",
    "  • Si 2 élèves ont le même numéro, ils sont liés au même parent",
    "",
    "Colonne H - WhatsApp (OPTIONNEL)",
    "  • Valeurs acceptées : Oui, Non, Yes, No, True, False, 1, 0",
    "  • Exemple : Oui",
    "",
    "⚠️ RÈGLES IMPORTANTES",
    "",
    "• Les colonnes Nom, Prénom et Classe sont OBLIGATOIRES",
    "• Un élève est considéré comme doublon si : même nom + même prénom + même classe",
    "• Les doublons ne seront pas importés (évite les répétitions)",
    "• Les parents existants sont reconnus par leur NUMÉRO DE TÉLÉPHONE",
    "• Si un parent avec le même téléphone existe, l'élève lui sera lié",
    "• Les espaces inutiles (début/fin) sont automatiquement supprimés",
    "",
    "✅ EXEMPLE D'IMPORT",
    "",
    "Vous importez :",
    "  - Jean Dupont (classe 6ème A) - Parent : Marie Dupont, +33612345678",
    "  - Sophie Dupont (classe 6ème A) - Parent : Marie Dupont, +33612345678",
    "",
    "Résultat après import :",
    "  - 2 élèves créés : Jean et Sophie",
    "  - 1 parent créé : Marie Dupont (avec 2 enfants liés)",
    "",
    "📧 CONTACT & SUPPORT",
    "",
    "En cas de problème lors de l'import :",
    "  1. Vérifiez que les noms de classe sont corrects (voir feuille 'Classes disponibles')",
    "  2. Vérifiez que les formats de date et téléphone sont exacts",
    "  3. Supprimez les lignes vides",
    "  4. Réessayez l'import",
]

for idx, instruction in enumerate(instructions, 2):
    cell = ws3.cell(row=idx, column=1)
    cell.value = instruction
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    if instruction.startswith("📋") or instruction.startswith("⚠️") or instruction.startswith("✅") or instruction.startswith("📧"):
        cell.font = Font(bold=True, size=12, color="003F7F")

# Sauvegarder le fichier
os.makedirs("uploads", exist_ok=True)
wb.save("uploads/exemple_import_eleves.xlsx")
print("✅ Fichier créé : uploads/exemple_import_eleves.xlsx")
