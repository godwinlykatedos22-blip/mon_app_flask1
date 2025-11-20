import sys, os
sys.path.insert(0, os.path.abspath(os.getcwd()))

from app import create_app
from models import db, Classe, Student, Parent
from openpyxl import Workbook
from datetime import date
import io

class TestConfig:
    SECRET_KEY = "test"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    WTF_CSRF_ENABLED = False
    TESTING = True
    LOGIN_DISABLED = True


def run_test():
    app = create_app(TestConfig)

    with app.app_context():
        db.create_all()

        # Créer une classe
        classe = Classe(name="6ème")
        db.session.add(classe)
        db.session.commit()

        # Créer un fichier Excel de test
        wb = Workbook()
        ws = wb.active
        ws.title = "Élèves"

        # En-têtes
        ws["A1"] = "Nom"
        ws["B1"] = "Prénom"
        ws["C1"] = "Date de naissance"
        ws["D1"] = "Classe"
        ws["E1"] = "Prénom Parent/Tuteur"
        ws["F1"] = "Nom Parent/Tuteur"
        ws["G1"] = "Téléphone Parent (E.164)"
        ws["H1"] = "WhatsApp (Oui/Non)"

        # Données
        ws["A2"] = "Dupont"
        ws["B2"] = "Jean"
        ws["C2"] = date(2008, 5, 15)
        ws["D2"] = "6ème"
        ws["E2"] = "Marie"
        ws["F2"] = "Dupont"
        ws["G2"] = "+33612345678"
        ws["H2"] = "Oui"

        ws["A3"] = "Dupont"
        ws["B3"] = "Sophie"
        ws["C3"] = date(2010, 3, 10)
        ws["D3"] = "6ème"
        ws["E3"] = "Marie"
        ws["F3"] = "Dupont"
        ws["G3"] = "+33612345678"
        ws["H3"] = "Oui"

        ws["A4"] = "Martin"
        ws["B4"] = "Alice"
        ws["C4"] = date(2009, 1, 20)
        ws["D4"] = "6ème"
        ws["E4"] = "Pierre"
        ws["F4"] = "Martin"
        ws["G4"] = "+33687654321"
        ws["H4"] = "Non"

        # Sauvegarder le fichier
        os.makedirs("uploads", exist_ok=True)
        file_path = "uploads/test_import.xlsx"
        wb.save(file_path)

        client = app.test_client()

        print("=" * 60)
        print("TEST: Import Excel avec Parents")
        print("=" * 60)

        # Uploader le fichier
        with open(file_path, 'rb') as f:
            data = {
                'excel_file': (f, 'test_import.xlsx')
            }
            resp = client.post('/eleves/import', data=data, follow_redirects=False)

        print(f"\n1️⃣ Import Excel : Status {resp.status_code}")
        assert resp.status_code == 302, "L'import devrait rediriger"

        # Vérifier les élèves importés
        students = Student.query.all()
        print(f"   ✓ Élèves importés : {len(students)}")
        assert len(students) == 3, "Devrait avoir 3 élèves"

        # Vérifier les parents
        parents = Parent.query.all()
        print(f"   ✓ Parents importés : {len(parents)}")
        assert len(parents) == 2, "Devrait avoir 2 parents (Marie et Pierre)"

        # Vérifier Marie Dupont (devrait avoir 2 enfants)
        marie = Parent.query.filter_by(last_name="Dupont").first()
        assert marie is not None, "Marie devrait exister"
        assert len(marie.students) == 2, "Marie devrait avoir 2 enfants"
        print(f"   ✓ Marie Dupont a {len(marie.students)} enfants liés")

        # Vérifier Pierre Martin (devrait avoir 1 enfant)
        pierre = Parent.query.filter_by(last_name="Martin").first()
        assert pierre is not None, "Pierre devrait exister"
        assert len(pierre.students) == 1, "Pierre devrait avoir 1 enfant"
        print(f"   ✓ Pierre Martin a {len(pierre.students)} enfant lié")

        # Vérifier WhatsApp
        assert marie.whatsapp_optin == True, "Marie devrait avoir WhatsApp activé"
        assert pierre.whatsapp_optin == False, "Pierre ne devrait pas avoir WhatsApp"
        print(f"   ✓ Statuts WhatsApp corrects")

        print("\n" + "=" * 60)
        print("✅ TOUS LES TESTS SONT PASSÉS !")
        print("=" * 60)

        # Nettoyer
        if os.path.exists(file_path):
            os.remove(file_path)


if __name__ == '__main__':
    run_test()
